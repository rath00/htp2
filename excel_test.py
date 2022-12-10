import os
import pathlib
import openpyxl


# 取得目前檔案所在的資料夾 
SRC_PATH =  pathlib.Path(__file__).parent.absolute()
UPLOAD_FOLDER = os.path.join(SRC_PATH)
print(UPLOAD_FOLDER)
#os.chdir('/content/drive/MyDrive/Colab Notebooks') # Colab 換路徑使用
wb = openpyxl.load_workbook(UPLOAD_FOLDER+'/HTP.xlsx', data_only=True)

names = wb.sheetnames    # 讀取 Excel 裡所有工作表名稱
s1 = wb['工作表1']        # 取得工作表名稱為「工作表1」的內容
s2 = wb.active           # 取得開啟試算表後立刻顯示的工作表 ( 範例為工作表 2 )
print(s1.title, s1.max_row, s1.max_column)
s1.insert_rows(2)
n=s1.max_row+1
s1.cell(2,1).value = 'test%d'%s1.max_row
s1.cell(2,2).value = 100
s1.cell(2,3).value = 100
s1.cell(2,4).value = 100
s1.cell(2,5).value = 100
s1.cell(2,6).value = 100
s1.cell(2,7).value = 100
s1.cell(2,8).value = 100
s1.cell(2,9).value = 100
s1.cell(2,10).value = 100
s1.cell(2,11).value = 100
s1.cell(2,12).value = 100
s1.cell(2,13).value = 100


wb.save('HTP.xlsx')