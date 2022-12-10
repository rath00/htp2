# 參考資料：https://www.i4k.xyz/article/weixin_44613063/108038602


import numpy as np
import time
import cv2
import os
import pathlib
import openpyxl #使用excel的檔案


# 取得目前檔案所在的資料夾 
SRC_PATH =  pathlib.Path(__file__).parent.absolute()
UPLOAD_FOLDER = os.path.join(SRC_PATH)

print(UPLOAD_FOLDER)


def detect(image_position):
    # 測試print(image_position)
    # 導入 YOLO 辨識類別：
    LABELS = open(UPLOAD_FOLDER+"/obj(1)/obj.names").read().strip().split("\n")
    np.random.seed(666)
    COLORS = np.random.randint(0, 255, size=(len(LABELS), 3), dtype="uint8")
    # 導入 YOLO 配置和權重文件並加載網路：
    net = cv2.dnn.readNetFromDarknet(UPLOAD_FOLDER+'/obj(1)/yolov4-tiny-obj.cfg', UPLOAD_FOLDER+'/obj(1)/yolov4-tiny-obj_best.weights')
    # 獲取 YOLO 未連接的輸出圖圖層
    layer = net.getUnconnectedOutLayersNames()
    image = cv2.imread(image_position)
    # 獲取圖片尺寸
    (H, W) = image.shape[:2]
    # 從輸入圖像構造一個blob，然后執行 YOLO 對象檢測器的前向傳遞，給我们邊界盒和相關概率
    blob = cv2.dnn.blobFromImage(image, 1/255.0, (416, 416),swapRB=True, crop=False)
    net.setInput(blob)
    start = time.time()
    # 前向傳遞，獲得信息
    layerOutputs = net.forward(layer)
    # 用於得出檢測時間
    end = time.time()
    print("YOLO took {:.6f} seconds".format(end - start))
    
    # 建立儲存不同資訊的陣列
    boxes = []
    confidences = []
    classIDs = []
    txtas=[]

    txt=''

    # 循環提取每個輸出層
    for output in layerOutputs:
        # 循環提取每個框
        for detection in output:
            # 提取當前目標的 類別ID(編號) 和 可信度
            scores = detection[5:]
            classID = np.argmax(scores)     # np.argmax(a) 找a陣列中的最大值
            confidence = scores[classID]

            # 通過確保檢測概率大於最小概率來過濾弱預測
            if confidence > 0.0:
                # 將邊界座標相對於圖像的大小進行縮放，YOLO 返回的是邊界框的中心(x,y)座標
                # 後面是邊界框的寬度和高度
                box = detection[0:4] * np.array([W, H, W, H])
                (centerX, centerY, width, height) = box.astype("int")
                txta=detection[0:4]
                (a, b, c, d) =txta.astype("float")

                # 轉換出邊框左上角座標
                x = int(centerX - (width / 2))
                y = int(centerY - (height / 2))

                # 更新邊界框座標、可信度和類ID的列表
                boxes.append([x, y, int(width), int(height)])
                confidences.append(float(confidence))
                classIDs.append(classID)
                txtas.append([a, b, c, d])
        
    # 非最大值抑制，確定唯一邊框，原始參數:idxs = cv2.dnn.NMSBoxes(boxes, confidences, 0.5, 0.3)
    idxs = cv2.dnn.NMSBoxes(boxes, confidences, 0.0, 0.3)
        
    # 建立紀錄偵測各個類別種數的陣列
    door=0
    window=0
    chimney=0
    crown=0
    bark=0
    fruit=0
    face=0
    body=0
    neck=0
    house=0
    tree=0
    person=0
    classes=[door,window,chimney,crown,bark,fruit,face,body,neck,house,tree,person]
    hps=[]
    tps=[]
    pps=[]
     # 設定房樹人整體位置的初始值
    (x2,y2)=(H, W)
    (w2, h2)=(0,0)

    # 確定每個對象只少有一個框存在
    if len(idxs) > 0:
        # 循環畫出表存的邊框
        for i in idxs.flatten():
            # 提取座標和寬度
            (x, y) = (boxes[i][0], boxes[i][1])
            (w, h) = (boxes[i][2], boxes[i][3])
            # 畫出邊框和標籤
            color = [int(c) for c in COLORS[classIDs[i]]]
            #  cv2.rectangle(影像, 頂點座標, 對向頂點座標, 顏色, 線條寬度, 線條類型)
            cv2.rectangle(image, (x, y), (x + w, y + h), color, 2, lineType=cv2.LINE_AA)
            text = "{}: {:.4f}".format(LABELS[classIDs[i]], confidences[i])
            #  cv2.putText(影像, 文字, 座標, 字型, 大小, 顏色, 線條寬度, 線條種類)
            cv2.putText(image, text, (x, y - 5), cv2.FONT_HERSHEY_SIMPLEX,
                1.0, color, 2, lineType=cv2.LINE_AA)
            # 記錄到個類別偵測到多少物件
            for j in range(len(classes)):
                if LABELS[classIDs[i]]==LABELS[j]:
                    classes[j]=classes[j]+1
            # 自動標籤
            txt=txt+'%d'%classIDs[i]+' %f'%txtas[i][0]+' %f'%txtas[i][1]+' %f'%txtas[i][2]+' %f'%txtas[i][3]+'\n'

            # 取得房樹人中心座標
            if LABELS[classIDs[i]]=='house':
                hp=[x+(w/2),y+(h/2)]
                hps.append(hp)
                print('House:')
                print(hps)
            elif LABELS[classIDs[i]]=='tree':
                tp=[x+(w/2),y+(h/2)]
                tps.append(tp)
                print('Tree:')
                print(tps)
            elif LABELS[classIDs[i]]=='person':
                pp=[x+(w/2),y+(h/2)]
                pps.append(pp)
                print('Person:')
                print(pps)
        

        #print('\n'+txt) # 測試自動標註的檔案
        # 儲存自動標註的.txt檔案
        name=os.path.basename(image_position)
        name = name.split(".")[0] #去掉檔案格式，取得檔案原名稱
        file = open(UPLOAD_FOLDER+'/txt/'+name+'.txt','w+')
        file.write(txt)
        file.close()
            
        # 找出房樹人整體的位置
        # 循環畫出表存的邊框
        for j in idxs.flatten():
            print('%d/'%boxes[j][0]+'%d/'%boxes[j][1]+'%d/'%boxes[j][2]+'%d/'%boxes[j][3])
            # 提取座標和寬度
            (x2, y2) = (min(x2,boxes[j][0] ), min(y2,boxes[j][1] ))
            (w2, h2) = (max(w2,boxes[j][0]+boxes[j][2] ), max(h2,boxes[j][1]+boxes[j][3] ))

        #print('\n'+'%d/'%x2+'%d/'%y2+'%d/'%w2+'%d/'%h2) #測試all的邊框   
        # 畫出邊框和標籤
        color = [int(c) for c in COLORS[classIDs[i]]]
        #  cv2.rectangle(影像, 頂點座標, 對向頂點座標, 顏色, 線條寬度, 線條類型)
        cv2.rectangle(image, (x2, y2), (w2, h2), (0,0,255), 2, lineType=cv2.LINE_AA)
        text = "{}: {:.4f}".format(LABELS[classIDs[i]], confidences[i])
        #  cv2.putText(影像, 文字, 座標, 字型, 大小, 顏色, 線條寬度, 線條種類)
        cv2.putText(image, 'all', (x2, y2 - 5), cv2.FONT_HERSHEY_SIMPLEX,
                1.0, (0,0,255), 2, lineType=cv2.LINE_AA)

    # 分析辨識結果，並產出圖片描述跟代表意義
    result=''
    result_dc=''
    warn=''
    # 儲存檔案名稱到excel中
    name=os.path.basename(image_position)
    name = name.split(".")[0] #去掉檔案格式，取得檔案原名稱
    wb = openpyxl.load_workbook(UPLOAD_FOLDER+'/HTP.xlsx', data_only=True) #開啟HTP.xlsx檔案
    s1 = wb['工作表1']        # 取得工作表名稱為「工作表1」的內容
    s2 = wb.active           # 取得開啟試算表後立刻顯示的工作表 ( 範例為工作表 2 )
    s1.insert_rows(2)
    s1.cell(2,1).value=name
    for i in range (len(classes)):
        s1.cell(2,i+2).value=classes[i]


    # classes=[0door,1window,2chimney,3crown,4bark,5fruit,6face,7body,8neck,9house,10tree,11person]
    # 確保有偵測到房樹人後再開始進行分析
    if classes[9]==0 and classes[10]==0 and classes[11]==0:
        result=result+'無法成功辨識，請再次輸入正確含有房樹人的圖片或是聯絡專業人士進行分析。'
        result_dc=result_dc+'此張圖系統無法成功辨識, 需再仔細閱讀和分析。'
        warn=warn+'無法成功進行辨識!'
        s1.cell(2,18).value='DETECT ERROR!'
    else:
        # 房的結果
        result=result+('  由你的圖可知，')
        if classes[0]>=1:   # 有門
            result=result+('你不排斥與外界接觸，') 
            result_dc=result_dc+('  在圖畫中出現了門，可以推測繪圖者可能比較外向，想要和他人接觸交流和被理解，也喜歡依賴他人；')
        elif classes[0]==0: # 沒門
            result=result+('你個性較獨立，')
            result_dc=result_dc+('  在圖畫中沒有發現門，推測繪圖者可能對外界比較有防禦心，'
            +'不喜歡他人主動靠近自己，個性也比較冷淡和退縮，在家庭方面可能有一些狀況，彼此之間比較缺乏互動、關係可能比較淡薄；') 
        if classes[1]>=1:   # 有窗
            result=result+('喜歡交朋友，')
            result_dc=result_dc+('在圖畫中出現了%d個窗戶'%classes[1]+'，可以得知繪圖者有想要和外界接觸交流的意願；')
        elif classes[1]==0: # 沒窗
            result=result+('比較謹慎，') 
            result_dc=result_dc+('在圖畫中沒有看到窗戶的存在，推測繪圖者可能比較退縮，和他人相處上也比較偏執，'+
            '另外還有可能有被害妄想的傾向，但是需要經過謹慎評估過後才能下定論，目前只是推測；') 
            warn=warn+'可能有被害妄想的傾向！'
        if classes[2]>=1:   # 有煙囪
            result=result+('重視家庭。') 
            result_dc=result_dc+('圖畫中出現了煙囪，表示繪圖者可能比較關注家庭和家庭給予的溫暖感。\n')
        elif classes[2]==0: # 沒煙囪
            result=result+('追求家庭溫暖。') 
            result_dc=result_dc+('圖畫中沒有煙囪，推測繪圖者最近可能比較消極，在家庭上缺乏心理上的溫暖。\n')
            warn=warn+'家庭關係可能較不好!'
        # 樹的結果 3crown,4bark,5fruit
        result=result+('另外，')
        if classes[3]>=1:   # 有樹冠
            result=result+('你較積極，') 
            result_dc=result_dc+('  在圖畫中有出現樹冠，推測繪圖者可能比較沒有自信上的問題，且有可能對自己有一定的自信和想要上進的心；')
        elif classes[3]==0: # 沒樹冠
            result=result+('你不熱衷交際，') 
            result_dc=result_dc+('  在圖畫中出現了沒有樹冠的樹，繪圖者可能比較自卑和抑鬱，個性比較內向不太能很好的與人交往，'
            +'另外還有患有思覺失調症的可能，但需要進一步了解，由專家診斷後才能確認；')
            warn=warn+'有思覺失調症的可能!'
        if classes[4]>=1:   # 有樹紋 
            result=result+('觀察力較敏銳') 
            result_dc=result_dc+('在圖畫中的樹上出現了樹皮，表示繪圖者在個性上可能對外界比較敏銳，在成長過程中可能有遇到一些事情，對他造成影響；')
            warn=warn+'成長過程可能有一些傷害!'
        elif classes[4]==0: # 沒樹紋
            result=result+('且較有活力') 
            result_dc=result_dc+('在圖畫中的樹上很單純沒有樹皮，表示繪圖者可能在過去沒有遇到什麼困難，或是已經從困難中走出來了，目前生活比較積極；')
        if classes[5]>=1:   # 有果實
            result=result+('，期待成功的未來。') 
            result_dc=result_dc+('圖中出現了%d個果實'%classes[5]+'，繪圖者可能有一些慾望和目標。\n')
        elif classes[5]==0: # 沒果實
            result=result+('。') 
            result_dc=result_dc+('圖中是單純的樹，樹上沒有果實，繪圖者可能目前沒有設立可實現的目標，對自己的評價也是，目前對自己可能沒有什麼要求。\n')
        # 人的結果 6face,7body,8neck
        result=result+('同時，')
        if classes[7]>=1:   # 有火柴人
            result=result 
            result_dc=result_dc+('  在畫作中出現了火柴人的身體，表示繪圖者具有掩飾性，說謊能力較強，可能不太願意表露真實的自我，有防禦性；')
        elif classes[7]==0: # 沒火柴人
            result=result 
            result_dc=result_dc+('  在畫作中沒有火柴人的身體，繪圖者對此測驗有一定的信任程度，對測驗比較沒有懷疑的心態，也不太有隱藏和掩飾；')
        if classes[6]>=1:   # 有五官
            result=result+('你容易適應環境，') 
            result_dc=result_dc+('在圖畫中有有五官的人，表示繪圖者可能願意和外界接觸，也比較沒有在新環境適應不良的問題；')
        elif classes[6]==0: # 沒五官
            result=result+('你個性較害羞，') 
            result_dc=result_dc+('在圖畫中出現了沒有五官的人，表示繪圖者可能偏好逃避人際關係，不太能很好的適應環境，個性比較害羞也比較強調自我。')
        if classes[8]>=1:   # 有脖子
            result=result+('較不易被情緒左右') 
            result_dc=result_dc+('圖畫中的人有脖子，表示繪圖者在智力與情緒之間是有連結的，比較不會只憑本能做事。\n\n')
        elif classes[8]==0: # 沒脖子
            result=result+('較易展現個人特質') 
            result_dc=result_dc+('圖畫中的人是沒有脖子的，可能表示繪圖者在智力與情緒之間沒有連結，做事可能會比較衝動，不顧後果，'
            +'在適應能力上可能也比較弱、乏靈活性。\n\n')
        
        
        # 對房樹整體大小和位置的分析
        # 距離關係
        Far=''
        Near=''
        print(len(hps))
        for i in range(len(hps)):
            hpx=hps[i][0]
            hpy=hps[i][1]
            for j in range(len(tps)):
                tpx=tps[j][0]
                tpy=tps[j][1]
                for k in range(len(pps)):
                    ppx=pps[k][0]
                    ppy=pps[k][1]
                    # 房和樹
                    ht=((hpx-tpx)**2+(hpy-tpy)**2)**(1/2)
                    # 樹和人
                    tp=((tpx-ppx)**2+(tpy-ppy)**2)**(1/2)
                    # 人和房
                    ph=((ppx-hpx)**2+(ppy-hpy)**2)**(1/2)
                    print('%d'%ht+',%d'%tp+',%d'%ph)

                    far=max(ht,tp,ph)
                    near=min(ht,tp,ph)
                    print('far:%d'%far+'near:%d'%near)

                    if far==ht:
                        Far='ht'
                    elif far==tp:
                        Far='tp'
                    elif far==ph:
                        Far='ph'
                    if near==ht:
                        Near='ht'
                    elif near==tp:
                        Near='tp'
                    elif near==ph:
                        Near='ph'
    
        # 畫面中距離最近的物件
        if Near=='ht':
            print('房子和樹離的最近')
            result=result+('。整體而言，你和家人感情良好')
            result_dc=result_dc+('  在圖畫中房子和樹之間的距離是最近的，'
            +'繪圖者可能與家人相處較密切')
            s1.cell(2,14).value = 'ht'
        elif Near=='tp':
            print('樹和人離的最近')
            result=result+('。整體而言，你和大家感情都不錯')
            result_dc=result_dc+('  在圖畫中樹和人之間的距離是最近的，'
            +'繪圖者與外界的關係良好')
            s1.cell(2,14).value = 'tp'
        elif Near=='ph':
            print('人和房子離的最近')
            result=result+('。整體而言，你喜歡你現在的家庭')
            result_dc=result_dc+('  在圖畫中人和房子之間的距離是最近的，'
            +'繪圖者可能喜歡他的家庭，且比較依賴家庭')
            s1.cell(2,14).value = 'ph'
        # 畫面中距離最遠的物件
        if Far=='ht':
            print('房子和樹離的最遠')
            result_dc=result_dc+('；在圖畫中房子和樹之間的距離是最遠的，'
            +'繪圖者可能與家人相處上較不密切。')
            s1.cell(2,15).value = 'ht'
        elif Far=='tp':
            print('樹和人離的最遠')
            result_dc=result_dc+('；在圖畫中樹和人之間的距離是最遠的，'
            +'繪圖者可能與外界比較疏離。')
            s1.cell(2,15).value = 'tp'
        elif Far=='ph':
            print('人和房子離的最遠')
            result_dc=result_dc+('；在圖畫中人和房子之間的距離是最遠的，'
            +'繪圖者可能不太喜歡現在的家庭關係，與家人之間比較疏離。')
            s1.cell(2,15).value = 'ph'
        if classes[9]>1 or classes[10]>1 or classes[11]>1:
            result_dc=result_dc+('但系統偵測到在圖畫中有%d個房子、'%classes[9]+'%d個樹'%classes[10]+'和%d個人，'%classes[11]+'彼此的互動關係可能需要再另外確認！')

        ax=w2-x2
        ay=h2-y2
        area_obj=ax*ay
        area_pic=H*W
        
        # 大小及位置關係
        if area_obj>(2*area_pic/3):
            print('畫面大於三分之二')
            result=result+('，是率直熱情的人。') # 原：而在個性上你可能比較善待自己、重視自身，在內心的情緒是比較高昂的，可能有好動、率直和易受情緒波動影響的特質。
            result_dc=result_dc+('\n'+'  在圖畫中，房、數、人三者的比例已經超過畫面的三分之二，繪圖者可能比較強調自我的存在，對環境和氣氛的變化比較沒有感覺，'
            +'但內心充滿緊張、幻想和敵意，比較具有侵略性、恐嚇性和攻擊性，也比較好動、情緒化跟率直。')
            s1.cell(2,16).value = '>2/3'
        elif area_obj<(area_pic/9):
            print('畫面小於九分之一')
            result=result+('，是內向但不討厭人群的人') # 原：而在個性上你可能比較內向， 不太擅長適應新環境，對自己也比較沒有信心，也比較依賴他人。
            result_dc=result_dc+('\n'+'  在圖畫中，房、數、人三者的比例已經小於畫面的九分之一，繪圖者可能比較害羞內向，不適應環境，'
            +'且對自我比較壓抑不太有自信，可能缺乏安全感，比較退縮和依賴別人，若有人打破他的自我意識時，可能就會顯得比較焦慮和沮喪。')
            s1.cell(2,16).value = '<1/9'
        else:
            s1.cell(2,16).value = 'Normal'
        
        if area_obj<=(area_pic/3):
            # print('圖畫比重小於三分之一了，要比較位置！') # 檢查和測試用   
            # 找中心位置
            mx=(x2+w2)/2
            my=(y2+h2)/2
            # 左側
            if 0<= mx< (W/3) and (H/3)<=my<=(2*H/3):
                print('圖形在左邊')
                result=result+('，且較容易懷念過去。') # 原：另外你可能也比較感性，在個性上因此比較容易衝動，另外也相對念舊，會留念過去。
                result_dc=result_dc+('\n'+'  在圖畫中，房、數、人三者的比例小於畫面的三分之一且集中在畫面左間，'
                +'左側象徵了過去、感情世界和女性化，這表示繪圖者在個性上可能比較衝動，且可能專注於過去，對過去有所留念。')
                s1.cell(2,17).value = 'left'
            # 中間
            elif (W/3)<= mx< (2*W/3) and (H/3)<=my<=(2*H/3):
                print('圖形在中間')
                result=result+('。') # 原：另外你有較強的自我意識，但在內心中可能又有一些不安，在努力想要維持心中的平衡。
                result_dc=result_dc+('\n'+'  在圖畫中，房、數、人三者的比例小於畫面的三分之一且集中在畫面中間，'
                +'表示了安全感，且自我意識可能較強，比較以自我為中心，若是成人的話，可能在內在有不安感，想要努力維持內心的平衡；'
                +'若是兒童的話，可能表示他比較在意自我，可塑性較差，不太能客觀的認識環境。')
                s1.cell(2,17).value = 'middle'
            # 右側
            elif (2*W/3)<= mx< W and (H/3)<=my<=(2*H/3):
                print('圖形在右邊')
                result=result+('，且對未來有憧憬。') # 原：另外你可能比較理性，比較能自我控制，對未來抱有憧憬。
                result_dc=result_dc+('\n'+'  在圖畫中，房、數、人三者的比例小於畫面的三分之一且集中在畫面右側，'
                +'右側象徵未來、理智和男性化，這表示繪圖者在個性上可能比較理性，且可能比較更關注未來而不是現在。')
                s1.cell(2,17).value = 'right'
            # 上側
            elif (W/3)<= mx< (2*W/3) and 0<=my<=(H/3):
                print('圖形在上側')
                result=result+('，且個性樂觀。') # 原：另外你可能正在追求過於遠大的目標、有比較多期望，個性樂觀也很有想法和想像力，但可能會高估自己的能力，'+'另外還可能會讓人覺得有點距離感、不好親近。'
                result_dc=result_dc+('\n'+'  在圖畫中，房、數、人三者的比例小於畫面的三分之一且集中在畫面上部，'
                +'表示繪圖者可能正在追求遠大的目標，個性過於樂觀、喜歡幻想，但自我期許太高也比較缺乏洞察力，有比較多和遠大的慾望，'
                +'另外可能還會給別人難以接近的距離感。')
                s1.cell(2,17).value = 'up'
            # 下側
            elif (W/3)<= mx< (2*W/3) and (2*H/3)<=my<=H:
                print('圖形在下側')
                result=result+('，且較喜歡在熟悉的環境。') # 原：另外你可能比較悲觀和注重現實，對自己比不太有自信，性格比較謹慎在熟悉的事物身邊才比較有安全感。
                result_dc=result_dc+('\n'+'  在圖畫中，房、數、人三者的比例小於畫面的三分之一且集中在畫面下部，'
                +'表示繪圖者可能比較沒有安全感，不太能適應，個性上也比較悲觀和注重現實，情緒有比較低落的傾向。')
                s1.cell(2,17).value = 'down'
            else:
                print('圖形在四個角的其中一邊')
                result=result+('，且較念舊。') # 原：另外你比較喜歡在人群中，比較有依賴性，個性也比較保守，偏好舊有的事物，不太喜歡嘗試陌生的東西。
                result_dc=result_dc+('\n'+'  在圖畫中，房、數、人三者的比例小於畫面的三分之一且集中在畫面角落，'
                +'表示繪圖者可能不夠有安全感跟自信，害怕獨立比較依賴他人，也可能不太喜歡嘗試新的事物，喜歡沈迷在幻想中。')
                s1.cell(2,17).value = 'corner'
        elif area_obj>(area_pic/3) and area_obj<=(2*area_pic/3):
            result=result+('。')
            s1.cell(2,17).value = 'Normal'

    # 測試產生的結果
    print('\n'+'\n'+result)
    print('\n'+result_dc)
    # 儲存產生的結果
    #name=os.path.basename(image_position)
    #name = name.split(".")[0] #去掉檔案格式，取得檔案原名稱

    file = open(UPLOAD_FOLDER+'/result/'+name+'-result.txt','w+')
    file.write(result)
    file.close()

    file = open(UPLOAD_FOLDER+'/result-dc/'+name+'-result-dc.txt','w+')
    file.write(warn+'/\n'+result_dc)
    file.close()
    
    # 儲存excel檔案   
    wb.save('HTP.xlsx')

    # 測試偵測結果（可產生圖片）
    # cv2.imshow(name, image)
    # cv2.waitKey(0)

# 定義讀取圖片的函式，供網頁使用
def return_img(img_local_path):
    """
    工具函數:
    獲取本地圖片流
    :param img_local_path:文件單張圖片的本地絕對路徑
    :return: 圖片流
    """
    import base64
    img_stream = ''
    with open(img_local_path, 'rb') as img_f:
        img_stream = img_f.read()
        img_stream = base64.b64encode(img_stream).decode()
    return img_stream

detect('/Users/wen/Downloads/HTP/htp-2/result-pic/2022-11-23-003716.png')#測試/Users/wen/Downloads/auto-label/picture/20220608-021.jpg

